import requests
from bs4 import BeautifulSoup
import openpyxl

def fetch_article_text(url):
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')

    # Extract title and text
    title = soup.find('h1').get_text()
    paragraphs = soup.find_all('p')
    article_text = ' '.join([para.get_text() for para in paragraphs])

    return title, article_text
def save_article_text(url_id, title, text):
    with open(f'{url_id}.txt', 'w', encoding='utf-8') as file:
        file.write(title + '\n' + text)
def main():
    # Load the input Excel file
    file_path = 'C:\\Internship\\20211030 Test Assignment-20240602T051347Z-001\\20211030 Test Assignment\\Input.xlsx'
    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        url_id, url = row
        title, text = fetch_article_text(url)
        save_article_text(url_id, title, text)

if __name__ == "__main__":
    main()
import os
import re
import nltk
from nltk.tokenize import word_tokenize, sent_tokenize
from textblob import TextBlob
import openpyxl

nltk.download('punkt')

# Load positive and negative words
positive_words = set(open('positive-words.txt').read().split())
negative_words = set(open('negative-words.txt').read().split())

def analyze_text(text):
    blob = TextBlob(text)
    
    sentences = sent_tokenize(text)
    words = word_tokenize(text)
    
    # Calculate scores
    positive_score = sum(1 for word in words if word in positive_words)
    negative_score = sum(1 for word in words if word in negative_words)
    polarity_score = blob.sentiment.polarity
    subjectivity_score = blob.sentiment.subjectivity
    avg_sentence_length = len(words) / len(sentences)
    complex_words_count = sum(1 for word in words if len(word) > 2)
    percentage_complex_words = (complex_words_count / len(words)) * 100
    fog_index = 0.4 * (avg_sentence_length + percentage_complex_words)
    avg_words_per_sentence = len(words) / len(sentences)
    syllable_per_word = sum(len(re.findall(r'[aeiouy]+', word.lower())) for word in words) / len(words)
    personal_pronouns = len(re.findall(r'\b(I|we|my|ours|us)\b', text, re.I))
    avg_word_length = sum(len(word) for word in words) / len(words)
    word_count = len(words)
    
    return {
        'positive_score': positive_score,
        'negative_score': negative_score,
        'polarity_score': polarity_score,
        'subjectivity_score': subjectivity_score,
        'avg_sentence_length': avg_sentence_length,
        'percentage_complex_words': percentage_complex_words,
        'fog_index': fog_index,
        'avg_words_per_sentence': avg_words_per_sentence,
        'complex_words_count': complex_words_count,
        'word_count': word_count,
        'syllable_per_word': syllable_per_word,
        'personal_pronouns': personal_pronouns,
        'avg_word_length': avg_word_length,
    }

def main():
    input_wb = openpyxl.load_workbook('Input.xlsx')
    input_sheet = input_wb.active

    output_wb = openpyxl.Workbook()
    output_sheet = output_wb.active
    output_sheet.append([
        "URL_ID", "URL", "POSITIVE SCORE", "NEGATIVE SCORE", "POLARITY SCORE",
        "SUBJECTIVITY SCORE", "AVG SENTENCE LENGTH", "PERCENTAGE OF COMPLEX WORDS",
        "FOG INDEX", "AVG NUMBER OF WORDS PER SENTENCE", "COMPLEX WORD COUNT",
        "WORD COUNT", "SYLLABLE PER WORD", "PERSONAL PRONOUNS", "AVG WORD LENGTH"
    ])

    for row in input_sheet.iter_rows(min_row=2, values_only=True):
        url_id, url = row
        with open(f'{url_id}.txt', 'r', encoding='utf-8') as file:
            text = file.read()
        
        analysis = analyze_text(text)
        output_sheet.append([
            url_id, url,
            analysis['positive_score'], analysis['negative_score'],
            analysis['polarity_score'], analysis['subjectivity_score'],
            analysis['avg_sentence_length'], analysis['percentage_complex_words'],
            analysis['fog_index'], analysis['avg_words_per_sentence'],
            analysis['complex_words_count'], analysis['word_count'],
            analysis['syllable_per_word'], analysis['personal_pronouns'],
            analysis['avg_word_length']
        ])
    
    output_wb.save('C:\\Internship\\20211030 Test Assignment-20240602T051347Z-001\\20211030 Test Assignment\\Output Data Structure.xlsx')

if __name__ == "__main__":
    main()
pip install beautifulsoup4 requests nltk textblob openpyxl
python extract_articles.py
python analyze_articles.py




